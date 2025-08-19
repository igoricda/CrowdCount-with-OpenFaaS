import pickle
import cv2
import requests
import json
import subprocess
import base64
import numpy as np
import re
import datetime
from requests.exceptions import Timeout, RequestException, ConnectionError
import openpyxl
import os
import glob # Used to find the RAPL file
from dotenv import load_dotenv
load_dotenv()

# Define the name for the output Excel file

XLSX_FILE = os.getenv("XLSX_FILE_TFLITE_SERVER_1R")

# --- Energy Measurement Functions (using RAPL interface) ---

def find_rapl_energy_file():
    """Finds the path to the RAPL energy file for the CPU package."""
    # Common path for Intel CPU package energy. It's usually the first one (e.g., intel-rapl:0)
    rapl_paths = glob.glob('/sys/class/powercap/intel-rapl:*/energy_uj')
    if not rapl_paths:
        # Fallback for AMD CPUs
        rapl_paths = glob.glob('/sys/class/power_cap/dram-*-*/energy')
        if not rapl_paths:
             return None
    return rapl_paths[0]


def get_rapl_energy(rapl_file_path):
    """
    Reads the current energy consumption from the RAPL file.
    The value in the file is in microjoules (uJ).
    This function converts it to milliwatt-hours (mWh) for consistency.
    1 mWh = 3,600,000 uJ
    """
    try:
        with open(rapl_file_path, 'r') as f:
            energy_uj = int(f.read())
            # Convert microjoules to milliwatt-hours
            energy_mwh = energy_uj / 3.6e6
            return energy_mwh
    except (IOError, ValueError) as e:
        print(f"Could not read RAPL energy file: {e}. Returning 0.")
        return 0

# === Excel (.xlsx) Setup ===
def setup_workbook(filename):
    """Loads an existing workbook or creates a new one if it doesn't exist."""
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Results"
        print(f"Created new workbook: {filename}")
    return workbook, sheet

def get_next_free_column(worksheet, row):
    """Finds the next available starting column in a given row."""
    col = 1
    # Check if the cell in the specified row and column has a value
    while worksheet.cell(row=row + 3, column=col).value is not None:
        col += 6  # Move to the next block of columns
    return col

def setup_openfaas():
    try:
        with open("/dev/null", "w") as nullfile:
            login_script = os.getenv("LOGIN_SCRIPT_SERVER")
            subprocess.run(["sudo", "/bin/bash", login_script],
                         check=True, stdout=nullfile, stderr=nullfile)
        print("OpenFaaS connection established successfully.")
        return True
    except subprocess.CalledProcessError:
        print("Error: Unable to connect to OpenFaaS server.")
        return False

def prepare_image(img_path):
    img = cv2.imread(img_path)
    if img is None:
        print(f"Error: Could not load image at {img_path}")
        return None
    try:
        return pickle.dumps(img)
    except Exception as e:
        print(f"Image processing error: {e}")
        return None

if __name__ == "__main__":
    # --- Setup RAPL energy monitoring ---
    rapl_path = find_rapl_energy_file()
    if rapl_path is None:
        print("Error: Could not find RAPL 'energy_uj' file.")
        print("Please ensure you are on Linux and have the 'intel-rapl' or 'amd_energy' kernel module loaded.")
        print("You may need to run this script with sudo.")
        exit(1)
    
    print(f"Using RAPL energy file for measurements: {rapl_path}")
    print("NOTE: This measures CPU package energy, not full system power.")

    image_list = [ "0p0f_0.jpg", "0p0f_1.jpg", "0p0f_2.jpg","0p0f_3.jpg", "0p0f_4.jpg",
                   "1p1f_0.jpg", "1p1f_1.jpg", "1p1f_2.jpg","1p1f_3.jpg", "1p1f_4.jpg",
                   "2p0f_0.jpg","2p1f_0.jpg","2p2f_0.jpg","2p2f_1.jpg", "2p2f_2.jpg",
                   "3p0f_0.jpg","3p2f_0.jpg","3p3f_0.jpg","3p3f_1.jpg","3p3f_2.jpg",
                   "4p1f_0.jpg","4p3f_0.jpg","4p3f_0.jpg", "4p3f_2.jpg", "4p4f_0.jpg",
                   "5p0f_0.jpg", "5p1f_0.jpg", "6p6f_0.jpg", "8p7f_0.jpg"]

    if not setup_openfaas():
        exit(1)

    # --- .xlsx File Handling ---
    workbook, sheet = setup_workbook(XLSX_FILE)
    url = os.getenv("OPENFAAS_URL_SERVER")
    openfaas_url = url + "/function/crowdcounttflite"

    # Logic to find the starting row and column
    start_row = 1
    while True:
        col = get_next_free_column(sheet, start_row)
        if col < 25:
            break
        start_row += 25

    for img_spec in image_list:
        directory = os.getenv("IMAGE_DIRECTORY")
        img_path = os.path.join(directory, img_spec)
        imdata = prepare_image(img_path)
        if imdata is None:
            continue

        json_data = json.dumps({
            "image_data": {
                "image": base64.b64encode(imdata).decode('ascii')
            }
        })
        
        # --- Write Headers to .xlsx ---
        sheet.cell(row=start_row + 1, column=col, value=f"Run {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        sheet.cell(row=start_row + 2, column=col, value=f"Image: {img_spec}")
        sheet.cell(row=start_row + 3, column=col, value="Iteration")
        sheet.cell(row=start_row + 3, column=col + 1, value="Count")
        sheet.cell(row=start_row + 3, column=col + 2, value="Elapsed Time (s)")
        sheet.cell(row=start_row + 3, column=col + 3, value="Energy Start (mWh)")
        sheet.cell(row=start_row + 3, column=col + 4, value="Energy End (mWh)")
        sheet.cell(row=start_row + 3, column=col + 5, value="Energy Request (mWh)")

        total_elapsed_time = 0
        times = []
        total_energy = 0
        energy = []

        n = 5
        iteration_count = 0
        for i in range(n):
            try:
                # Get energy reading right before the request
                energy_before = get_rapl_energy(rapl_path)
                
                sheet.cell(row=start_row + 4 + i, column=col + 3, value=energy_before)

                response = requests.post(openfaas_url, data=json_data, timeout=300, headers={'Content-Type': 'application/json'})
                response.raise_for_status()

                # Get energy reading right after the request
                energy_after = get_rapl_energy(rapl_path)

                try:
                    result = response.json()
                except json.JSONDecodeError:
                    matches = re.findall(rb'({.*})', response.content)
                    if matches:
                        result = json.loads(matches[-1].decode())
                    else:
                        print("Error: No JSON object found in response.")
                        print("Raw response:", response.content)
                        continue

                elapsed_time = response.elapsed.total_seconds()
                total_elapsed_time += elapsed_time
                times.append(elapsed_time)
                
                energy_request = energy_after - energy_before
                # RAPL counter can sometimes wrap around (reset to 0). Handle this case.
                if energy_request < 0:
                    print("Warning: RAPL energy counter may have reset. Discarding this reading.")
                    # You could try to read the 'max_energy_range_uj' file to handle this perfectly,
                    # but for short-lived requests, skipping the value is often sufficient.
                    continue
                
                total_energy += energy_request
                energy.append(energy_request)
                
                print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] Image: {img_spec} | Iteration: {i+1}")
                print(f"Count: {result['count']}, Time: {elapsed_time:.4f}s, Energy: {energy_request:.6f} mWh")

                # --- Write Iteration Data to .xlsx ---
                sheet.cell(row=start_row + 4 + i, column=col, value=i + 1)
                sheet.cell(row=start_row + 4 + i, column=col + 1, value=result['count'])
                sheet.cell(row=start_row + 4 + i, column=col + 2, value=elapsed_time)
                sheet.cell(row=start_row + 4 + i, column=col + 4, value=energy_after)
                sheet.cell(row=start_row + 4 + i, column=col + 5, value=energy_request)
                iteration_count += 1

            except Timeout:
                print(f"Error for image {img_spec}: Request timed out after 300 seconds")
            except ConnectionError as e:
                print(f"Connection error for image {img_spec}: {e}")
            except RequestException as e:
                print(f"Request failed for image {img_spec}: {e}")
            except Exception as e:
                print(f"An unexpected error occurred for image {img_spec}: {e}")

        # --- Write Summary Stats to .xlsx ---
        summary_start_row = 12
        sheet.cell(row = start_row + summary_start_row-1, column = col, value = "Summary")
        sheet.cell(row = start_row + summary_start_row, column = col, value = "Total Time")
        sheet.cell(row = start_row + summary_start_row, column = col+1, value = total_elapsed_time)
        sheet.cell(row = start_row + summary_start_row + 1, column = col, value = "Average Time")
        sheet.cell(row = start_row + summary_start_row + 1, column = col+1, value = total_elapsed_time / len(times))
        sheet.cell(row = start_row + summary_start_row + 2, column = col, value = "Variance")
        sheet.cell(row = start_row + summary_start_row + 2, column = col+1, value = np.var(times))
        sheet.cell(row = start_row + summary_start_row + 3, column = col, value = "Std Dev")
        sheet.cell(row = start_row + summary_start_row + 3, column = col+1, value = np.std(times))
        sheet.cell(row = start_row + summary_start_row + 4, column = col, value = "Min Time")
        sheet.cell(row = start_row + summary_start_row + 4, column = col+1, value = min(times))
        sheet.cell(row = start_row + summary_start_row + 5, column = col, value = "Max Time")
        sheet.cell(row = start_row + summary_start_row + 5, column = col+1, value = max(times))
        sheet.cell(row = start_row + summary_start_row + 6, column = col, value = "Total Requests")
        sheet.cell(row = start_row + summary_start_row + 6, column = col+1, value = i)
        sheet.cell(row = start_row + summary_start_row + 7, column = col, value = "Total Energy")
        sheet.cell(row = start_row + summary_start_row + 7, column = col+1, value = total_energy)
        sheet.cell(row = start_row + summary_start_row + 8, column = col, value = "Average Energy")
        sheet.cell(row = start_row + summary_start_row + 8, column = col+1, value = total_energy / len(energy))
        sheet.cell(row = start_row + summary_start_row + 9, column = col, value = "Variance Energy")
        sheet.cell(row = start_row + summary_start_row + 9, column = col+1, value = np.var(energy))
        sheet.cell(row = start_row + summary_start_row + 10, column = col, value = "Std Dev Energy")
        sheet.cell(row = start_row + summary_start_row + 10, column = col+1, value = np.std(energy))

        # Save the workbook after each image's data is fully written
        workbook.save(XLSX_FILE)
        print(f"Data for {img_spec} saved to {XLSX_FILE}\n")
        
        # --- Update column and row for the next run ---
        col += 6
        if col > 24:
            col = 1
            start_row += 25
            
    print("All images processed. Final results saved.")