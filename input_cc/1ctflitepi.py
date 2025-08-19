import pickle
import cv2
import requests
import json
import subprocess
import base64
import numpy as np
import re
import datetime
from requests.exceptions import Timeout, RequestException, ConnectionError
import openpyxl  # Replaced gspread
import os        # Added for file operations
import serial
import threading
import time
from dotenv import load_dotenv
load_dotenv()

# Define the name for the output Excel file
XLSX_FILE = os.getenv("XLSX_FILE_TFLITE_PI_1R")

def read_serial_and_compute_energy(shared_data, data_lock, port='/dev/ttyUSB0', baudrate=115200):
    try:
        with serial.Serial(port, baudrate, timeout=1) as ser:
            prev_time = None
            shared_data["total_mWh"] = 0.0
            while True:
                line = ser.readline().decode('utf-8', errors='ignore').strip()
                if not line:
                    continue
                try:
                    timestamp_str, current_str, voltage_str = line.split(";")
                    timestamp = int(timestamp_str)
                    current = float(current_str)
                    voltage = float(voltage_str)
                    power = current * voltage  # mW
                    if prev_time is not None:
                        delta_ms = timestamp - prev_time
                        energy_mWh = (power * delta_ms) / 3600000
                        with data_lock:
                            shared_data["total_mWh"] += energy_mWh
                    prev_time = timestamp
                except ValueError:
                    continue
    except serial.SerialException as e:
        print(f"Serial error: {e}")
    except KeyboardInterrupt:
        print("\nStopped by user.")

# === Excel (.xlsx) Setup ===
def setup_workbook(filename):
    """Loads an existing workbook or creates a new one if it doesn't exist."""
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Results"
        print(f"Created new workbook: {filename}")
    return workbook, sheet

def get_next_free_column(worksheet, row):
    """Finds the next available starting column in a given row."""
    col = 1
    while worksheet.cell(row=row + 3, column=col).value is not None:
        col += 6
    return col

def setup_openfaas():
    login_script = os.getenv("LOGIN_SCRIPT_RASPBERRYPI")  
    try:
        with open("/dev/null", "w") as nullfile:
            subprocess.run(["sudo", "/bin/bash", login_script],
                         check=True, stdout=nullfile, stderr=nullfile)
        print("OpenFaaS connection established successfully.")
        return True
    except subprocess.CalledProcessError:
        print("Error: Unable to connect to OpenFaaS server.")
        return False

def prepare_image(img_path):
    img = cv2.imread(img_path)
    if img is None:
        print(f"Error: Could not load image at {img_path}")
        return None
    try:
        return pickle.dumps(img)
    except Exception as e:
        print(f"Image processing error: {e}")
        return None

if __name__ == "__main__":
    shared_data = {"total_mWh": 0.0}
    data_lock = threading.Lock()

    serial_thread = threading.Thread(
        target=read_serial_and_compute_energy,
        args=(shared_data, data_lock),
        daemon=True
    )
    serial_thread.start()
    
    image_list = [ "0p0f_0.jpg", "0p0f_1.jpg", "0p0f_2.jpg","0p0f_3.jpg", "0p0f_4.jpg",
                   "1p1f_0.jpg", "1p1f_1.jpg", "1p1f_2.jpg","1p1f_3.jpg", "1p1f_4.jpg",
                   "2p0f_0.jpg","2p1f_0.jpg","2p2f_0.jpg","2p2f_1.jpg", "2p2f_2.jpg",
                   "3p0f_0.jpg","3p2f_0.jpg","3p3f_0.jpg","3p3f_1.jpg","3p3f_2.jpg",
                   "4p1f_0.jpg","4p3f_0.jpg","4p3f_0.jpg", "4p3f_2.jpg", "4p4f_0.jpg",
                   "5p0f_0.jpg", "5p1f_0.jpg", "6p6f_0.jpg", "8p7f_0.jpg"]

    if not setup_openfaas():
        exit(1)

    # --- .xlsx File Handling ---
    workbook, sheet = setup_workbook(XLSX_FILE)
    url = os.getenv("OPENFAAS_URL_RASPERRYPI")
    openfaas_url = url + "/function/crowdcounttflite"

    # Logic to find the starting row and column
    start_row = 1
    while True:
        col = get_next_free_column(sheet, start_row)
        if col < 25:
            break
        start_row += 25

    for img_spec in image_list:
        directory = os.getenv("IMAGE_DIRECTORY")
        img_path = os.path.join(directory, img_spec)
        imdata = prepare_image(img_path)
        if imdata is None:
            continue

        json_data = json.dumps({
            "image_data": {
                "image": base64.b64encode(imdata).decode('ascii')
            }
        })

        # --- Write Headers to .xlsx ---
        sheet.cell(row=start_row + 1, column=col, value=f"Run {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        sheet.cell(row=start_row + 2, column=col, value=f"Image: {img_spec}")
        sheet.cell(row=start_row + 3, column=col, value="Iteration")
        sheet.cell(row=start_row + 3, column=col + 1, value="Count")
        sheet.cell(row=start_row + 3, column=col + 2, value="Elapsed Time")
        sheet.cell(row=start_row + 3, column=col + 3, value="Energy Start")
        sheet.cell(row=start_row + 3, column=col + 4, value="Energy End")
        sheet.cell(row=start_row + 3, column=col + 5, value="Energy Request")

        total_elapsed_time = 0
        times = []
        total_energy = 0
        energy = []
        iteration_count = 0

        n = 5
        for i in range(n):
            try:
                with data_lock:
                    energy_before = shared_data["total_mWh"]
                
                response = requests.post(openfaas_url, data=json_data, timeout=300, headers={'Content-Type': 'application/json'})
                response.raise_for_status()
                
                with data_lock:
                    energy_after = shared_data["total_mWh"]

                try:
                    result = response.json()
                except json.JSONDecodeError:
                    matches = re.findall(rb'({.*})', response.content)
                    if matches:
                        result = json.loads(matches[-1].decode())
                    else:
                        print("Error: No JSON object found in response. Raw response:", response.content)
                        continue

                elapsed_time = response.elapsed.total_seconds()
                total_elapsed_time += elapsed_time
                times.append(elapsed_time)
                
                energy_request = energy_after - energy_before
                total_energy += energy_request
                energy.append(energy_request)
                iteration_count += 1
                
                print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] Image: {img_spec} | Iteration: {i+1}")
                print(f"Count: {result['count']}, Time: {elapsed_time:.4f}s, Energy: {energy_request:.6f} mWh")

                # --- Write Iteration Data to .xlsx ---
                sheet.cell(row=start_row + 4 + i, column=col, value=i + 1)
                sheet.cell(row=start_row + 4 + i, column=col + 1, value=result['count'])
                sheet.cell(row=start_row + 4 + i, column=col + 2, value=elapsed_time)
                sheet.cell(row=start_row + 4 + i, column=col + 3, value=energy_before)
                sheet.cell(row=start_row + 4 + i, column=col + 4, value=energy_after)
                sheet.cell(row=start_row + 4 + i, column=col + 5, value=energy_request)

            except (Timeout, ConnectionError, RequestException) as e:
                print(f"An error occurred during request for {img_spec}: {e}")
            except Exception as e:
                print(f"An unexpected error occurred for {img_spec}: {e}")

        # --- Write Summary Stats if any iterations were successful ---
        if times and energy:
            summary_start_row = start_row + 4 + n + 2
            sheet.cell(row=summary_start_row - 1, column=col, value="Summary")
            sheet.cell(row=summary_start_row, column=col, value="Total Time")
            sheet.cell(row=summary_start_row, column=col + 1, value=total_elapsed_time)
            sheet.cell(row=summary_start_row + 1, column=col, value="Average Time")
            sheet.cell(row=summary_start_row + 1, column=col + 1, value=np.mean(times))
            sheet.cell(row=summary_start_row + 2, column=col, value="Variance")
            sheet.cell(row=summary_start_row + 2, column=col + 1, value=np.var(times))
            sheet.cell(row=summary_start_row + 3, column=col, value="Std Dev")
            sheet.cell(row=summary_start_row + 3, column=col + 1, value=np.std(times))
            sheet.cell(row=summary_start_row + 4, column=col, value="Min Time")
            sheet.cell(row=summary_start_row + 4, column=col + 1, value=min(times))
            sheet.cell(row=summary_start_row + 5, column=col, value="Max Time")
            sheet.cell(row=summary_start_row + 5, column=col + 1, value=max(times))
            sheet.cell(row=summary_start_row + 6, column=col, value="Total Requests")
            sheet.cell(row=summary_start_row + 6, column=col + 1, value=iteration_count)
            sheet.cell(row=summary_start_row + 7, column=col, value="Total Energy")
            sheet.cell(row=summary_start_row + 7, column=col + 1, value=total_energy)
            sheet.cell(row=summary_start_row + 8, column=col, value="Average Energy")
            sheet.cell(row=summary_start_row + 8, column=col + 1, value=np.mean(energy))
            sheet.cell(row=summary_start_row + 9, column=col, value="Variance Energy")
            sheet.cell(row=summary_start_row + 9, column=col + 1, value=np.var(energy))
            sheet.cell(row=summary_start_row + 10, column=col, value="Std Dev Energy")
            sheet.cell(row=summary_start_row + 10, column=col + 1, value=np.std(energy))

        # Save the workbook and move to the next column block
        workbook.save(XLSX_FILE)
        print(f"Data for {img_spec} saved to {XLSX_FILE}\n")
        col += 6
        if col > 24:
            col = 1
            start_row += 25
            
    print("All images processed. Final results saved.")