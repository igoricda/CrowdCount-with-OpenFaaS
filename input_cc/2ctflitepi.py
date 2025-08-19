import pickle
import cv2
import requests
import json
import subprocess
import base64
import numpy as np
import re
import datetime
from requests.exceptions import Timeout, RequestException, ConnectionError
import openpyxl  # Replaced gspread
import os        # Added for file operations
import serial
import threading
import time
import concurrent.futures
from dotenv import load_dotenv
load_dotenv()

# Define the name for the output Excel file
XLSX_FILE = os.getenv("XLSX_FILE_TFLITE_PI_2R")

def read_serial_and_compute_energy(shared_data, data_lock, port='/dev/ttyUSB0', baudrate=115200):
    try:
        with serial.Serial(port, baudrate, timeout=1) as ser:
            #print(f"Connected to {port} at {baudrate} baud.")
            #print("Time (ms) | Current (mA) | Voltage (V) | Power (mW) | Energy (mWh)")
            #print("-" * 70)

            prev_time = None
            shared_data["total_mWh"] = 0.0
            while True:
                line = ser.readline().decode('utf-8', errors='ignore').strip()
                if not line:
                    continue

                try:
                    timestamp_str, current_str, voltage_str = line.split(";")
                    timestamp = int(timestamp_str)
                    current = float(current_str)
                    voltage = float(voltage_str)
                    power = current * voltage  # mW

                    if prev_time is not None:
                        delta_ms = timestamp - prev_time
                        energy_mWh = (power * delta_ms) / 3600000  # mWh
                        with data_lock:
                            shared_data["total_mWh"] += energy_mWh

                    else:
                        delta_ms = 0
                        energy_mWh = 0

                    prev_time = timestamp

                    #print(f"{timestamp} | {current:.2f} mA | {voltage:.2f} V | {power:.2f} mW | {total_mWh:.6f} mWh")

                except ValueError:
                    continue

    except serial.SerialException as e:
        print(f"Serial error: {e}")
    except KeyboardInterrupt:
        print("\nStopped by user.")

# === Excel (.xlsx) Setup (Replaces Google Sheets setup) ===
def setup_workbook(filename):
    """Loads an existing workbook or creates a new one."""
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Results"
        print(f"Created new workbook: {filename}")
    return workbook, sheet

def get_next_free_column(worksheet, row):
    col = 1
    while worksheet.cell(row=row+3, column=col).value:
        col += 6
    return col

def setup_openfaas():
    try:
        with open("/dev/null", "w") as nullfile:
            login_script = os.getenv("LOGIN_SCRIPT_RASPBERRYPI")
            subprocess.run(["sudo", "/bin/bash", login_script],
                         check=True, stdout=nullfile, stderr=nullfile)
        print("OpenFaaS connection established successfully.")
        return True
    except subprocess.CalledProcessError:
        print("Error: Unable to connect to OpenFaaS server.")
        return False

def prepare_image(img_path):
    img = cv2.imread(img_path)
    if img is None:
        print(f"Error: Could not load image at {img_path}")
        return None
    try:
        return pickle.dumps(img)
    except Exception as e:
        print(f"Image processing error: {e}")
        return None

def request(shared_data, data_lock, json_data, openfaas_url, current_energy):

        response = requests.post(openfaas_url, data=json_data, timeout=300, headers={'Content-Type': 'application/json'})
        response.raise_for_status()
        try:
                result = response.json()
        except json.JSONDecodeError:
                    matches = re.findall(rb'({.*})', response.content)
                    if matches:
                        result = json.loads(matches[-1].decode())
                    else:
                        print("Error: No JSON object found in response.")
                        print("Raw response:", response.content)
                        exit(1)

        with data_lock:
                    energy_after = shared_data["total_mWh"]
        elapsed_time = response.elapsed.total_seconds()
        print(f"Response time: {elapsed_time:.2f} seconds")
        print(f"Count: {result['count']}, Energy Start: {current_energy}, Energy End: {energy_after}")

        return result['count'], elapsed_time, energy_after


if __name__ == "__main__":
    shared_data = {"total_mWh": 0.0}
    data_lock = threading.Lock()
    serial_thread = threading.Thread(
        target=read_serial_and_compute_energy,
        args=(shared_data, data_lock),
        daemon=True
    )
    serial_thread.start()

    image_list = [ "0p0f_0.jpg", "0p0f_1.jpg", "0p0f_2.jpg","0p0f_3.jpg", "0p0f_4.jpg", \
                   "1p1f_0.jpg", "1p1f_1.jpg", "1p1f_2.jpg","1p1f_3.jpg", "1p1f_4.jpg", \
                   "2p0f_0.jpg","2p1f_0.jpg","2p2f_0.jpg","2p2f_1.jpg", "2p2f_2.jpg", \
                   "3p0f_0.jpg","3p2f_0.jpg","3p3f_0.jpg","3p3f_1.jpg","3p3f_2.jpg", \
                   "4p1f_0.jpg","4p3f_0.jpg","4p3f_0.jpg", "4p3f_2.jpg", "4p4f_0.jpg", \
                   "5p0f_0.jpg", "5p1f_0.jpg", "6p6f_0.jpg", "8p7f_0.jpg", \
                    ]

    if not setup_openfaas():
        exit(1)
        
    # --- .xlsx File Handling ---
    workbook, sheet = setup_workbook(XLSX_FILE)
    url = os.getenv("OPENFAAS_URL_RASPERRYPI")
    openfaas_url = url + "/function/crowdcounttflite"

    start_row = 1
    while(True):
        try:
            col = get_next_free_column(sheet, start_row)
            if col >= 25:
                raise IndexError
            break
        except IndexError:
            if start_row == 1:
                start_row += 24
            else:
                start_row += 25
            continue

    for img_spec in image_list:
        directory = os.getenv("IMAGE_DIRECTORY")
        img_path = os.path.join(directory, img_spec)

        imdata = prepare_image(img_path)
        if imdata is None:
            exit(1)

        json_data = json.dumps({
            "image_data": {
                "image": base64.b64encode(imdata).decode('ascii')
            }
        })

        # --- Sheet Header Updates using openpyxl ---
        sheet.cell(row=start_row, column=col, value=f"Run {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        sheet.cell(row=start_row+1, column=col, value=f"Image: {img_spec}")
        sheet.cell(row=start_row + 2, column=col, value="Iteration")
        sheet.cell(row=start_row + 2, column=col + 1, value="Count1 and Count2")
        sheet.cell(row=start_row + 2, column=col + 2, value="Elapsed Time")
        sheet.cell(row=start_row + 2, column=col + 3, value="Energy Start")
        sheet.cell(row=start_row + 2, column=col + 4, value="Energy End")
        sheet.cell(row=start_row + 2, column=col + 5, value="Energy Request")

        total_elapsed_time = 0
        times = []
        total_energy = 0
        energy = []
        count1l = []
        count2l = []
        request_energy = []
        for i in range (5):
            with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
                with data_lock:
                    current_energy = shared_data["total_mWh"]
                
                # --- Sheet Update using openpyxl ---
                sheet.cell(row=start_row+3+i, column=col + 3, value=current_energy)
                
                print("Starting iteration", i+1, "for", img_spec)
                future1 = executor.submit(request, shared_data, data_lock, json_data, openfaas_url, current_energy)
                future2 = executor.submit(request, shared_data, data_lock, json_data, openfaas_url, current_energy)
                count1, time1, energy_request1 = future1.result()
                count2, time2, energy_request2 = future2.result()
                count1l.append((count1))
                count2l.append((count2))
                times.append(max(time1, time2))
                energy.append(max(energy_request1, energy_request2))
                total_elapsed_time += times[i]
                print(f"Iteration {i+1} for {img_spec} completed. Count1: {count1}, Count2: {count2}, Time: {times[i]}s, Energy Request: {energy[i]}mWh")
                req_energy = energy[i] - current_energy
                request_energy.append(req_energy if req_energy > 0 else 0)
                total_energy += request_energy[i]

            # --- Sheet Updates ---
            sheet.cell(row=start_row + 3 + i, column=col, value=i + 1)
            sheet.cell(row=start_row + 3 + i, column=col + 1 , value=str(count1l[i]) + ", " + str(count2l[i]) )
            sheet.cell(row=start_row + 3 + i, column=col + 2, value=times[i])
            sheet.cell(row=start_row + 3 + i, column=col + 4 , value=energy[i])
            sheet.cell(row=start_row + 3 + i, column=col + 5, value=request_energy[i])
            print(f"Iteration {i+1} for {img_spec} completed and saved. Time: {times[i]}s, Energy: {energy[i]}mWh")

        # --- Summary stats ---
        summary_start_row = 4 + 5 + 3
        sheet.cell(row=start_row + summary_start_row-1, column=col, value="Summary")
        sheet.cell(row=start_row + summary_start_row, column=col, value="Total Time")
        sheet.cell(row=start_row + summary_start_row, column=col+1, value=total_elapsed_time)
        sheet.cell(row=start_row + summary_start_row + 1, column=col, value="Average Time")
        sheet.cell(row=start_row + summary_start_row + 1, column=col+1, value=total_elapsed_time / (len(times)*2))
        sheet.cell(row=start_row + summary_start_row + 2, column=col, value="Variance")
        sheet.cell(row=start_row + summary_start_row + 2, column=col+1, value=np.var(times))
        sheet.cell(row=start_row + summary_start_row + 3, column=col, value="Std Dev")
        sheet.cell(row=start_row + summary_start_row + 3, column=col+1, value=np.std(times))
        sheet.cell(row=start_row + summary_start_row + 4, column=col, value="Min Time")
        sheet.cell(row=start_row + summary_start_row + 4, column=col+1, value=min(times))
        sheet.cell(row=start_row + summary_start_row + 5, column=col, value="Max Time")
        sheet.cell(row=start_row + summary_start_row + 5, column=col+1, value=max(times))
        sheet.cell(row=start_row + summary_start_row + 6, column=col, value="Total Requests")
        sheet.cell(row=start_row + summary_start_row + 6, column=col+1, value=(i+1)*2)
        sheet.cell(row=start_row + summary_start_row + 7, column=col, value="Total Energy")
        sheet.cell(row=start_row + summary_start_row + 7, column=col+1, value=total_energy)
        sheet.cell(row=start_row + summary_start_row + 8, column=col, value="Average Energy")
        sheet.cell(row=start_row + summary_start_row + 8, column=col+1, value=total_energy / (len(energy)*2))
        sheet.cell(row=start_row + summary_start_row + 9, column=col, value="Variance Energy")
        sheet.cell(row=start_row + summary_start_row + 9, column=col+1, value=np.var(request_energy))
        sheet.cell(row=start_row + summary_start_row + 10, column=col, value="Std Dev Energy")
        sheet.cell(row=start_row + summary_start_row + 10, column=col+1, value=np.std(request_energy))
        
        # Save the workbook after each image's data is fully written
        workbook.save(XLSX_FILE)

        col += 6
        if col >= 25:
            col = 1
            if start_row == 1:
                start_row += 24
            else:
                start_row += 25


        print(f"Finished processing for {img_spec}. Waiting before next image...")
        time.sleep(1) # Retaining a small delay